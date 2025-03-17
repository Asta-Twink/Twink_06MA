# ___StandAlone INIT___
from Env import *
if Mod_Work == True:SAR = True
else: SAR = False

@Exception_Handle
def Register_FN(Rgtr):
    #_________ Defintion_List _________
    globals()["Blob_Image_Nomi"] = None
    globals()["Blob_Image_Sign"] = None
    globals()["Blob_Image_Emp"] = None

    @Exception_Handle
    def Temp_View_pwd():
        pwd.Pwd_txt.clear()
        pwd.Pwd_txt.setEchoMode(QLineEdit.Password)
        Rgtr.NonPFTOPF.setVisible(True)
        if Rgtr.Temp_View.isChecked() == True:
            @Exception_Handle
            def Pwd_Connect():
                pwd.close()
                if pwd.Pwd_txt.text() == "SIL":
                    globals()["Blob_Image_Nomi"] = None
                    globals()["Blob_Image_Sign"] = None
                    globals()["Blob_Image_Emp"] = None
                    Customer_List = DB_Fetch(
                                             "select emp_code,team,employee_name,blood_group,phone_no from "
                                             "register where active = 'Y'  and  emp_code like 'TEMP%' order by UID",
                                             False, "LOL")

                    Push_Table_Values(Rgtr.OQTB_EmployeeList, Customer_List, False)
                    @Exception_Handle
                    def NonPF_To_PF_Convert():
                        table_list=[]

                        new_id = Emp_code_Gen(True)
                        try:

                            sql = fr"UPDATE register SET emp_code = '{new_id}' WHERE UID = '{globals()['product_list'][0]}' and emp_code = '{globals()['Product_Master_Data']}'";
                            DB_Cmt(sql,False)
                            output = []
                            myresult = DB_Fetch("Show Tables",False,"LOE")
                            for step in myresult:

                                try:
                                    a = int(step.split("_")[1])
                                    if a == 2023:
                                        output.append(step)

                                except:
                                    pass
                            output.append('advance_details')
                            for step in output:
                                try:
                                    sql = fr"UPDATE {step} SET empcode = '{new_id}' WHERE empcode = '{globals()['Product_Master_Data']}'";
                                    DB_Cmt( sql, False)
                                except Exception as e:
                                    print(e)
                                    pass

                            Customer_List = DB_Fetch(
                                                     "select emp_code,team,employee_name,blood_group,"
                                                     "phone_no from "
                                                     "register where active = 'Y'  and  Eemp_code like 'TEMP%'order by "
                                                     "UID",
                                                     False, "LOL")
                            Push_Table_Values(Rgtr.OQTB_EmployeeList, Customer_List, False)
                            Adjust_Table_Width(Rgtr.OQTB_EmployeeList, [100, 300, 75, 75, 100])

                        except Exception as e:
                            print(e)
                            print("hello","imnot working")
                    Rgtr.NonPFTOPF.clicked.connect(lambda: NonPF_To_PF_Convert())

            pwd.show()
            pwd.OK_Btn.clicked.connect(lambda: Pwd_Connect())
        else:
            Customer_List = DB_Fetch(
                                     "select emp_code,team,employee_name,blood_group,phone_no from register "
                                     "where active = 'Y'  and  emp_code like 'SIL%' order by UID",
                                     False, "LOL")

            Push_Table_Values(Rgtr.OQTB_EmployeeList, Customer_List, False)

        # if Rgtr.Temp_View.isChecked() == True:
        #     pwd.close()
        #     if pwd.Pwd_txt.text() == "AstA":
        #         Rgtr.NonPFTOPF.setVisible(True)
        #
        #         def NonPF_To_PF_Convert():
        #             print("Nonto")
        #             pass
        #
        #
        #
        #
        #     Rgtr.NonPFTOPF.clicked.connect(lambda: NonPF_To_PF_Convert())

    # Fetch_Table_individual_details
    @Exception_Handle
    def Fetch_tbl(row, column):
        Rgtr.GB_UpdateEmp.setChecked(False)
        Rgtr.GB_Add_Emp.setChecked(False)
        Rgtr.OQCB_Dept.clear()
        Rgtr.OQCB_BloodGroup.clear()
        Rgtr.OQCB_Team.clear()
        Rgtr.OQL_EmpPhoto.clear()
        Rgtr.OQL_SignPhoto.clear()
        Rgtr.OQL_NomPhoto.clear()
        try:
            Cus_index = Fetch_Table_Values(Rgtr.OQTB_EmployeeList)
            Ind_detail = DB_Fetch( f"select * from register where emp_code ='{Cus_index[row][0]}'", False, "LOE")
            print(Ind_detail)
            Rgtr.IQLE_EmpName.setText(Ind_detail[1])
            Rgtr.IQLE_EmpCode.setText(Ind_detail[2])
            Rgtr.OQCB_Dept.addItem(Ind_detail[3])
            Rgtr.IQLE_ESIC.setText(Ind_detail[4])
            Rgtr.IQLE_UAN.setText(Ind_detail[5])
            Rgtr.IQLE_PAN.setText(Ind_detail[6])
            Rgtr.IQLE_Aadhar.setText(Ind_detail[7])
            Rgtr.IQTE_Address.setText(Ind_detail[8])

            if Ind_detail[9] == 'Yes':
                Rgtr.IQRB_MrdYes.setChecked(True)
                Rgtr.IQRB_MrdNo.setChecked(False)
            else:
                Rgtr.IQRB_MrdNo.setChecked(True)
                Rgtr.IQRB_MrdYes.setChecked(False)
            Rgtr.IQLE_FatherSpouse.setText(Ind_detail[10])
            if Ind_detail[11] == 'M':
                Rgtr.IQRB_Male.setChecked(True)
            elif Ind_detail[11] == 'F':
                Rgtr.IQRB_Female.setChecked(True)
            else:
                Rgtr.IQRB_Other.setChecked(True)
            if Ind_detail[12] == 'Yes':
                Rgtr.IQRB_ShiftwrkYes.setChecked(True)
                Rgtr.IQRB_ShiftWrkNo.setChecked(False)
            else:
                Rgtr.IQRB_ShiftWrkNo.setChecked(True)
                Rgtr.IQRB_ShiftwrkYes.setChecked(False)
            Rgtr.IQLE_Shift1Salary.setText(str(Ind_detail[13]))
            Rgtr.IQLEShift2Salary.setText(str(Ind_detail[14]))
            Rgtr.IQLE_Shift3Salary.setText(str(Ind_detail[15]))
            Rgtr.IQLE_PhoneNo.setText(str(Ind_detail[16]))
            Rgtr.OQCB_BloodGroup.addItem(str(Ind_detail[17]))
            Rgtr.IQLE_BankAcNo.setText(str(Ind_detail[18]))
            Rgtr.IQLE_BankName.setText(str(Ind_detail[19]))
            Rgtr.IQLE_IFSC.setText(str(Ind_detail[20]))
            Rgtr.IQLE_Branch.setText(str(Ind_detail[21]))
            DoB = str(Ind_detail[22])
            DoJ = str(Ind_detail[23])
            Rgtr.OQDE_DOB.setDate(QDate.fromString(DoB, 'yyyy-MM-dd'))
            Rgtr.OQDE_DOJ.setDate(QDate.fromString(DoJ, 'yyyy-MM-dd'))
            Rgtr.IQLE_NomineeName.setText(str(Ind_detail[27]))
            Rgtr.IQLE_NomineePhNo.setText(str(Ind_detail[28]))
            # if Ind_detail[30] == 'PF':
            #     Rgtr.IQRB_PF.setChecked(True)
            #     Rgtr.IQRB_NonPF.setChecked(False)
            # else:
            #     Rgtr.IQRB_NonPF.setChecked(True)
            #     Rgtr.IQRB_PF.setChecked(False)
            if Ind_detail[30] == 'Y':
                Rgtr.OQC_OfficeStaff.setChecked(True)
            else:
                Rgtr.OQC_OfficeStaff.setChecked(False)
            Rgtr.IQCB_Worker.setEnabled(True)
            Rgtr.IQCB_Worker.setChecked(False)
            print(Ind_detail[33])
            if Ind_detail[33] == 'Y':
                Rgtr.IQCB_Worker.setChecked(True)
            else:
                Rgtr.IQCB_Worker.setChecked(False)

            Rgtr.IQCB_Worker.setEnabled(False)
            Rgtr.OQC_OfficeStaff.setEnabled(False)
            Rgtr.OQCB_Team.addItem(Ind_detail[31])

            if Ind_detail[25] != bytes(b'...'):
                try:
                    pixmap = get_pixmap_from_blob(Ind_detail[25])
                    Rgtr.OQL_EmpPhoto.setPixmap(pixmap)
                    Rgtr.OQL_EmpPhoto.setScaledContents(True)
                except:
                    Rgtr.OQL_EmpPhoto.clear()
                try:
                    pixmap_Sig = get_pixmap_from_blob(Ind_detail[26])
                    Rgtr.OQL_SignPhoto.setPixmap(pixmap_Sig)
                    Rgtr.OQL_SignPhoto.setScaledContents(True)
                except:
                    Rgtr.OQL_SignPhoto.clear()
                try:
                    pixmap_Nomi = get_pixmap_from_blob(Ind_detail[29])
                    Rgtr.OQL_NomPhoto.setPixmap(pixmap_Nomi)
                    Rgtr.OQL_NomPhoto.setScaledContents(True)
                except:
                    Rgtr.OQL_NomPhoto.clear()
            globals()["Product_Master_Data"] = Cus_index[row][0]
            globals()["product_list"]=Ind_detail
            Update_Property(Rgtr.NEX_EmployeeDetails, "Readonly", "IQLE", True)
            Update_Property(Rgtr.NEX_EmployeeDetails, "Enable", "IQRB", False)

        except Exception as e:
            traceback.print_exc()

    # Update_values
    @Exception_Handle
    def Enable_Update():
        try:
            if Rgtr.GB_UpdateEmp.isChecked():
                Update_Property(Rgtr.NEX_EmployeeDetails,"Readonly","IQLE",False)
                Update_Property(Rgtr.NEX_EmployeeDetails, "Enable", "IQRB", True)
                Rgtr.OQC_OfficeStaff.setEnabled(True)
                Rgtr.IQCB_Worker.setEnabled(True)
            else:
                Update_Property(Rgtr.NEX_EmployeeDetails,"Readonly","IQLE",True)
                Update_Property(Rgtr.NEX_EmployeeDetails, "Enable", "IQRB", False)
                Rgtr.OQC_OfficeStaff.setEnabled(False)
                Rgtr.IQCB_Worker.setEnabled(False)
        except Exception as e:
            print(e)
        Rgtr.OQCB_Dept.setEnabled(True)
        Rgtr.OQCB_Team.setEnabled(True)
        Rgtr.IQPB_PhotoBrowse.setEnabled(True)
        Rgtr.OQL_Sign.setEnabled(True)
        Rgtr.IQPB_NomPhotoBrowse.setEnabled(True)
        Rgtr.OQCB_BloodGroup.setEnabled(True)
        Rgtr.GB_Add_Emp.setChecked(False)
        Rgtr.OQCB_Team.addItems(Team)
        Rgtr.OQCB_Dept.addItems(Dept)
        Rgtr.OQCB_BloodGroup.addItems(Blod_grp)

    @Exception_Handle
    def Enable_Add():
        Rgtr.GB_UpdateEmp.setChecked(False)
        Rgtr.IQLE_EmpCode.clear()
        Rgtr.IQLE_EmpName.clear()
        Rgtr.OQCB_Dept.clear()
        Rgtr.IQLE_ESIC.clear()
        Rgtr.IQLE_UAN.clear()
        Rgtr.IQLE_PAN.clear()
        Rgtr.IQLE_Aadhar.clear()
        Rgtr.IQTE_Address.clear()

        Rgtr.IQRB_MrdYes.setChecked(False)
        Rgtr.IQRB_MrdNo.setChecked(False)
        Rgtr.IQRB_Male.setChecked(False)
        Rgtr.IQRB_Female.setChecked(False)
        Rgtr.IQRB_Other.setChecked(False)
        Rgtr.IQRB_ShiftwrkYes.setChecked(False)
        Rgtr.IQRB_ShiftWrkNo.setChecked(False)
        Rgtr.IQRB_PF.setEnabled(True)
        Rgtr.IQRB_NonPF.setEnabled(True)
        Rgtr.IQCB_Worker.setEnabled(True)
        Rgtr.IQRB_PF.setChecked(False)
        Rgtr.IQRB_NonPF.setChecked(False)
        Rgtr.OQC_OfficeStaff.setChecked(False)
        Rgtr.IQCB_Worker.setChecked(False)


        Rgtr.IQLE_FatherSpouse.clear()
        Rgtr.IQLE_Shift1Salary.clear()
        Rgtr.IQLEShift2Salary.clear()
        Rgtr.IQLE_Shift3Salary.clear()
        Rgtr.IQLE_PhoneNo.clear()
        Rgtr.OQCB_BloodGroup.clear()
        Rgtr.IQLE_BankAcNo.clear()
        Rgtr.IQLE_BankName.clear()
        Rgtr.IQLE_IFSC.clear()
        Rgtr.IQLE_Branch.clear()
        Rgtr.OQDE_DOB.setDate(QtCore.QDate.currentDate())
        Rgtr.OQDE_DOJ.setDate(QtCore.QDate.currentDate())
        Rgtr.IQLE_NomineeName.clear()
        Rgtr.IQLE_NomineePhNo.clear()
        Rgtr.OQCB_Team.clear()
        Rgtr.OQL_EmpPhoto.clear()
        Rgtr.OQL_SignPhoto.clear()
        Rgtr.OQL_NomPhoto.clear()

       #-----------------------------------
        Rgtr.OQCB_Dept.addItems(Dept)
        Rgtr.OQCB_Team.addItems(Team)
        Rgtr.OQCB_BloodGroup.addItems(Blod_grp)

        Rgtr.IQLE_EmpCode.setText(Emp_code_Gen(True))

        Rgtr.OQL_EmpPhoto.setText("")
        Rgtr.OQL_SignPhoto.setText("")
        Rgtr.OQL_NomPhoto.setText("")



        # Rgtr.IQRB_MrdYes.setEnabled(True)
        # Rgtr.IQRB_MrdNo.setEnabled(True)
        # Rgtr.IQRB_Male.setEnabled(True)
        # Rgtr.IQRB_Female.setEnabled(True)
        # Rgtr.IQRB_Other.setEnabled(True)
        # Rgtr.IQRB_ShiftwrkYes.setEnabled(True)
        # Rgtr.IQRB_ShiftWrkNo.setEnabled(True)

        Rgtr.OQC_OfficeStaff.setEnabled(True)

        Update_Property(Rgtr.NEX_EmployeeDetails,"Readonly","IQLE",False)
        Update_Property(Rgtr.NEX_EmployeeDetails, "Readonly", "IQTE", False)
        Update_Property(Rgtr.NEX_EmployeeDetails, "Enable", "IQRB", True)
        Update_Property(Rgtr.NEX_EmployeeDetails, "Readonly", "OQDE", False)
        # Rgtr.IQLE_EmpName.setReadOnly(False)
        # Rgtr.IQLE_EmpCode.setEnabled(False)
        # Rgtr.IQLE_ESIC.setReadOnly(False)
        # Rgtr.IQLE_UAN.setReadOnly(False)
        # Rgtr.IQLE_PAN.setReadOnly(False)
        # Rgtr.IQLE_Aadhar.setReadOnly(False)
        # Rgtr.IQTE_Address.setReadOnly(False)
        # Rgtr.IQLE_FatherSpouse.setReadOnly(False)
        # Rgtr.IQLE_Shift1Salary.setReadOnly(False)
        # Rgtr.IQLEShift2Salary.setReadOnly(False)
        # Rgtr.IQLE_Shift3Salary.setReadOnly(False)
        # Rgtr.IQLE_PhoneNo.setReadOnly(False)
        # Rgtr.IQLE_BankAcNo.setReadOnly(False)
        # Rgtr.IQLE_BankName.setReadOnly(False)
        # Rgtr.IQLE_IFSC.setReadOnly(False)
        # Rgtr.IQLE_Branch.setReadOnly(False)
        # Rgtr.OQDE_DOB.setReadOnly(False)
        # Rgtr.OQDE_DOJ.setReadOnly(False)
        # Rgtr.IQLE_NomineeName.setReadOnly(False)
        # Rgtr.IQLE_NomineePhNo.setReadOnly(False)

        Rgtr.IQPB_PhotoBrowse.setEnabled(True)
        Rgtr.OQL_Sign.setEnabled(True)
        Rgtr.IQPB_NomPhotoBrowse.setEnabled(True)

    @Exception_Handle
    def Update_Emp_Details():
        if UI_Confirmation(UI_Confirm_Win, fr"Please confirm do you want to Update the selected data'"):

            if globals()["Blob_Image_Nomi"] != None:
                cursor = db.cursor()
                insert_query_Nomi = "UPDATE register SET nominee_photo = %s WHERE emp_code = %s"
                data_Nomi = (globals()["Blob_Image_Nomi"], globals()["Product_Master_Data"])
                cursor.execute(insert_query_Nomi, data_Nomi)
                db.commit()
                cursor.close()


            if globals()["Blob_Image_Sign"] != None:
                insert_query_Sign = "UPDATE register SET signature = %s WHERE emp_code = %s"
                data_Sign = (globals()["Blob_Image_Sign"], globals()["Product_Master_Data"])
                cursor = db.cursor()
                cursor.execute(insert_query_Sign, data_Sign)
                db.commit()
                cursor.close()
                 


            if globals()["Blob_Image_Emp"] != None:
                cursor = db.cursor()
                insert_query_Emp = "UPDATE register SET photo = %s WHERE emp_code = %s"
                data_Emp = (globals()["Blob_Image_Emp"], globals()["Product_Master_Data"])
                cursor.execute(insert_query_Emp, data_Emp)
                db.commit()
                cursor.close()
                 


            DOB_Date = Rgtr.OQDE_DOB.date().toString(Qt.ISODate)
            DOJ_Date = Rgtr.OQDE_DOJ.date().toString(Qt.ISODate)


            if Rgtr.IQRB_MrdYes.isChecked():
                Mar = 'Yes'
            else:
                Mar = 'No'
            if Rgtr.IQRB_ShiftwrkYes.isChecked():
                Shift = 'Yes'
            else:
                Shift = 'No'
            if Rgtr.OQC_OfficeStaff.isChecked():
                Office = 'Y'
            else:
                Office = 'N'
            # if Rgtr.IQRB_PF.isChecked():
            #     PF = 'PF'
            # else:
            #     PF = 'Non PF'
            if Rgtr.IQRB_Male.isChecked():
                Gnd = "M"
            elif Rgtr.IQRB_Female.isChecked():
                Gnd = "F"
            else:
                Gnd = "O"
            if Rgtr.IQCB_Worker.isChecked():
                Work = 'Y'
            else:
                Work = 'N'

            dict = {
                'employee_name': Rgtr.IQLE_EmpName.text(),
                'designation': Rgtr.OQCB_Dept.currentText(),
                'esic_no': Rgtr.IQLE_ESIC.text(),
                'uan_no': Rgtr.IQLE_UAN.text(),
                'pan_no': Rgtr.IQLE_PAN.text(),
                'aadhar_no': Rgtr.IQLE_Aadhar.text(),
                'address': Rgtr.IQTE_Address.toPlainText(),
                'marriage_status': Mar,
                'f_sp_name': Rgtr.IQLE_FatherSpouse.text(),
                'gender': Gnd,
                'shift_work': Shift,
                'shift_1_salary': Rgtr.IQLE_Shift1Salary.text(),
                'shift_2_salary': Rgtr.IQLEShift2Salary.text(),
                'shift_3_salary': Rgtr.IQLE_Shift3Salary.text(),
                'phone_no': Rgtr.IQLE_PhoneNo.text(),
                'blood_group': Rgtr.OQCB_BloodGroup.currentText(),
                'bank_account_no': Rgtr.IQLE_BankAcNo.text(),
                'bank_name': Rgtr.IQLE_BankName.text(),
                'ifsc_code': Rgtr.IQLE_IFSC.text(),
                'branch': Rgtr.IQLE_Branch.text(),
                'date_of_birth': DOB_Date,
                'date_of_join': DOJ_Date,
                # 'photo' : globals()["Blob_Image"],
                # 'signature' : Rgtr.OQL_SignPhoto.pixmap(),
                # 'nominee_photo' : Rgtr.OQL_NomPhoto.pixmap(),
                'nominee_name': Rgtr.IQLE_NomineeName.text(),
                'nominee_phone_no': Rgtr.IQLE_NomineePhNo.text(),
                # 'ET': PF,
                'Worker':Work,
                'office_staff': Office,
                'team': Rgtr.OQCB_Team.currentText(),
            }
            c_dict = {
                'emp_code': globals()["Product_Master_Data"],
            }
            DB_Update_Dict(dict, c_dict, "register", False)
            Customer_List = DB_Fetch(
                                     "select emp_code,team,employee_name,blood_group,phone_no from register "
                                     "where active = 'Y' and emp_code like 'SIL%'order by UID",
                                     False, "LOL")
            Push_Table_Values(Rgtr.OQTB_EmployeeList, Customer_List, False)
            globals()["Blob_Image_Nomi"] = None
            globals()["Blob_Image_Sign"] = None
            globals()["Blob_Image_Emp"] = None
            UI_Confirmation(UI_Confirm_Win, "Updated Successfully")

    @Exception_Handle
    def Add_Emp_Details():
        DOB_Date = Rgtr.OQDE_DOB.date().toString(Qt.ISODate)
        DOJ_Date = Rgtr.OQDE_DOJ.date().toString(Qt.ISODate)
        DOB_Date = Rgtr.OQDE_DOB.date().toString(Qt.ISODate)
        DOJ_Date = Rgtr.OQDE_DOJ.date().toString(Qt.ISODate)
        if UI_Confirmation(UI_Confirm_Win, fr"Please confirm do you want to Add the Above data'"):
            if Rgtr.IQRB_MrdYes.isChecked():
                Mar = 'Yes'
            else:
                Mar = 'No'
            if Rgtr.IQRB_ShiftwrkYes.isChecked():
                Shift = 'Yes'
            else:
                Shift = 'No'
            if Rgtr.OQC_OfficeStaff.isChecked():
                Office = 'Y'
            else:
                Office = 'N'
            # if Rgtr.IQRB_PF.isChecked():
            #     PF = 'PF'
            # else:
            #     PF = 'Non PF'
            if Rgtr.IQRB_Male.isChecked():
                Gnd = "M"
            elif Rgtr.IQRB_Female.isChecked():
                Gnd = "F"
            else:
                Gnd = "O"
            if Rgtr.IQCB_Worker.isChecked():
                Work = 'Y'
            else:
                Work = 'N'
            dict = {
                'employee_name': Rgtr.IQLE_EmpName.text(),
                'emp_code': Rgtr.IQLE_EmpCode.text(),
                'designation': Rgtr.OQCB_Dept.currentText(),
                'esic_no': Rgtr.IQLE_ESIC.text(),
                'uan_no': Rgtr.IQLE_UAN.text(),
                'pan_no': Rgtr.IQLE_PAN.text(),
                'aadhar_no': Rgtr.IQLE_Aadhar.text(),
                'address': Rgtr.IQTE_Address.toPlainText(),
                'marriage_status': Mar,
                'f_sp_name': Rgtr.IQLE_FatherSpouse.text(),
                'gender': Gnd,
                'shift_work': Shift,
                'shift_1_salary': 0.0 if Rgtr.IQLE_Shift1Salary.text() == '' else Rgtr.IQLE_Shift1Salary.text(),
                'shift_2_salary': 0.0 if Rgtr.IQLEShift2Salary.text()== '' else Rgtr.IQLEShift2Salary.text(),
                'shift_3_salary': 0.0 if Rgtr.IQLE_Shift3Salary.text() == '' else Rgtr.IQLE_Shift3Salary.text(),
                'phone_no': Rgtr.IQLE_PhoneNo.text(),
                'blood_group': Rgtr.OQCB_BloodGroup.currentText(),
                'bank_account_no': Rgtr.IQLE_BankAcNo.text(),
                'bank_name': Rgtr.IQLE_BankName.text(),
                'ifsc_code': Rgtr.IQLE_IFSC.text(),
                'branch': Rgtr.IQLE_Branch.text(),
                'date_of_birth': DOB_Date,
                'date_of_join': DOJ_Date,
                'nominee_name': Rgtr.IQLE_NomineeName.text(),
                'nominee_phone_no': Rgtr.IQLE_NomineePhNo.text(),
                # 'ET': PF,
                'Worker': Work,
                'office_staff': Office,
                'team': Rgtr.OQCB_Team.currentText(),
            }
            DB_Push_Dict( dict, "register", False)

            Customer_List = DB_Fetch(
                                     "select emp_code,team,employee_name,blood_group,phone_no from register where"
                                     " active = 'Y' and emp_code like 'SIL%' order by UID",
                                     False, "LOL")
            Push_Table_Values(Rgtr.OQTB_EmployeeList, Customer_List, False)
            DB_Creation(Cur_Date_NF, False)

            Table_Data_Temp = DB_Fetch('select emp_code,team,employee_name,designation from register where '
                                       'active = "Y" order by team DESC,employee_name', False, 'LOL')
            for step in Table_Data_Temp:
                step.append(QCheckBox())
                QS = QSpinBox();
                QS.setRange(1, 3);
                step.append(QS)
                QS = QSpinBox();
                QS.setRange(0, 8);
                step.append(QS)
            Push_Table_Values(AttnPush.OQTB_Register, Table_Data_Temp, False)
            View_Table_Data = Attendance_datasplit(Attendance_Fetch(Cur_Date_MY), "Atn+ot", False)
            AttnView.OQTB_EmpDetails.horizontalHeader().setSortIndicator(0, Qt.AscendingOrder)
            Push_Table_Values(AttnView.OQTB_EmpDetails, View_Table_Data[0], False)
            Push_Table_Values(AttnView.OQTB_EmpAttendance, View_Table_Data[1], False)


            if globals()["Blob_Image_Nomi"] != None:
                cursor = db.cursor()
                insert_query_Nomi = "UPDATE register SET nominee_photo = %s WHERE emp_code = %s"
                data_Nomi = (globals()["Blob_Image_Nomi"], Rgtr.IQLE_EmpCode.text())
                cursor.execute(insert_query_Nomi, data_Nomi)
                db.commit()
                cursor.close()
                 


            if globals()["Blob_Image_Sign"] != None:
                insert_query_Sign = "UPDATE register SET signature = %s WHERE emp_code = %s"
                data_Sign = (globals()["Blob_Image_Sign"], Rgtr.IQLE_EmpCode.text())
                cursor = db.cursor()
                cursor.execute(insert_query_Sign, data_Sign)
                db.commit()
                cursor.close()
                 

            if globals()["Blob_Image_Emp"] != None:
                cursor = db.cursor()
                insert_query_Emp = "UPDATE register SET photo = %s WHERE emp_code = %s"
                data_Emp = (globals()["Blob_Image_Emp"], Rgtr.IQLE_EmpCode.text())
                cursor.execute(insert_query_Emp, data_Emp)
                db.commit()
                cursor.close()
                 
            globals()["Blob_Image_Nomi"] = None
            globals()["Blob_Image_Sign"] = None
            globals()["Blob_Image_Emp"] = None
            if UI_Confirmation(UI_Confirm_Win, "Employee Added Successfully"):
                # Clearing the values after successful addition
                Rgtr.IQLE_EmpName.setText("")
                Rgtr.IQLE_EmpCode.setText("")
                # Rgtr.OQCB_Dept.setText("")
                Rgtr.IQLE_ESIC.setText("")
                Rgtr.IQLE_UAN.setText("")
                Rgtr.IQLE_PAN.setText("")
                Rgtr.IQLE_Aadhar.setText("")
                Rgtr.IQTE_Address.setText("")
                Rgtr.IQLE_FatherSpouse.setText("")
                Rgtr.IQLE_Shift1Salary.setText("")
                Rgtr.IQLEShift2Salary.setText("")
                Rgtr.IQLE_Shift3Salary.setText("")
                Rgtr.IQLE_PhoneNo.setText("")
                # Rgtr.OQCB_BloodGroup.setText("")
                Rgtr.IQLE_BankAcNo.setText("")
                Rgtr.IQLE_BankName.setText("")
                Rgtr.IQLE_IFSC.setText("")
                Rgtr.IQLE_Branch.setText("")
                Rgtr.IQLE_NomineeName.setText("")
                Rgtr.IQLE_NomineePhNo.setText("")
                # Rgtr.OQCB_Team.setText("")
                Rgtr.IQLE_EmpCode.setText(Emp_code_Gen(True))

    @Exception_Handle
    def Right_Click_Menu(pos):
        row = Rgtr.OQTB_EmployeeList.selectedIndexes()
        R_menu = QMenu()
        Remove = QAction("Remove", R_menu)
        Revert = QAction("Revert", R_menu)
        R_menu.addAction(Remove)
        R_menu.addAction(Revert)
        pos = Rgtr.OQTB_EmployeeList.mapToGlobal(pos)
        Remove.triggered.connect(lambda: Delete_Employee(row[0].row()))
        Revert.triggered.connect(lambda: Revert_Employee())
        R_menu.exec_(pos)

    @Exception_Handle
    def Delete_Employee(index):
        dict = {
            'active': "N"
        }
        Employee = Fetch_Table_Values(Rgtr.OQTB_EmployeeList)[index][0]

        c_dict = {
            'emp_code': Employee
        }

        if UI_Confirmation(UI_Confirm_Win, f"Please confirm to delete the Employee :< {Employee} >:"):
            DB_Update_Dict( dict, c_dict, 'register', True)
            Customer_List = DB_Fetch("select emp_code,team,employee_name,blood_group,phone_no from register where "
                                     "active = 'Y'  and  emp_code like 'SIL%' order by UID",
                                     False, "LOL")
            Push_Table_Values(Rgtr.OQTB_EmployeeList, Customer_List, False)

    @Exception_Handle
    def Revert_DB_Push():
        temp = Rvt.OQTB_EmpRevert.selectedItems()
        if UI_Confirmation(UI_Confirm_Win, f"Please confirm to revert the Selected Employee"):
            for step in temp:
                dict = {
                    'active': "Y"
                }
                Employee = Fetch_Table_Values(Rvt.OQTB_EmpRevert)[step.row()][1]

                c_dict = {
                    'UID': Fetch_Employee_ID(Employee)
                }

                DB_Update_Dict(dict, c_dict, 'register', False)
            Push_Table_Values(Rvt.OQTB_EmpRevert,
                              DB_Fetch(
                                  'select emp_code,employee_name,designation,phone_no from register where active = "N"',
                                  False, "LOL"), False)
        Customer_List = DB_Fetch("select emp_code,team,employee_name,blood_group,phone_no from register where "
                                 "active = 'Y'  and  emp_code like 'SIL%' order by UID",
                                 False, "LOL")
        Push_Table_Values(Rgtr.OQTB_EmployeeList, Customer_List, False)

    Rvt.IQDB_EmpRevert.accepted.connect(lambda: Revert_DB_Push())
    Rvt.IQDB_EmpRevert.rejected.connect(lambda: Rvt.close())
    @Exception_Handle
    def Revert_Employee():
            Adjust_Table_Width(Rvt.OQTB_EmpRevert,(10,20,20,12))
            Push_Table_Values(Rvt.OQTB_EmpRevert,
                              DB_Fetch(
                                       'select emp_code,employee_name,designation,phone_no from register where active = "N"',
                                       False, "LOL"), False)
            Rvt.show()


    @Exception_Handle
    def read_image(file_path):
        with open(file_path, 'rb') as file:
            return file.read()

    @Exception_Handle
    def open_file_dialog_Emp_Img():
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        file_dialog = QFileDialog()
        file_dialog.setOptions(options)
        file_dialog.setNameFilter("Image Files (*.png *.jpg *.bmp)")
        file_dialog.setWindowTitle("Select an Image")
        file_dialog.setFileMode(QFileDialog.ExistingFile)
        if file_dialog.exec_() == QFileDialog.Accepted:
            selected_file = file_dialog.selectedFiles()[0]
            original_image_data = read_image(selected_file)
            new_width, new_height = 500, 300
            resized_image_data = resize_image(original_image_data, new_width, new_height)
            pixmap = QPixmap()
            pixmap.loadFromData(resized_image_data)
            Rgtr.OQL_EmpPhoto.setPixmap(pixmap)
            globals()["Blob_Image_Emp"] = resized_image_data

    @Exception_Handle
    def open_file_dialog_Emp_Sign():
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        file_dialog = QFileDialog()
        file_dialog.setOptions(options)
        file_dialog.setNameFilter("Image Files (*.png *.jpg *.bmp)")
        file_dialog.setWindowTitle("Select an Image")
        file_dialog.setFileMode(QFileDialog.ExistingFile)
        if file_dialog.exec_() == QFileDialog.Accepted:
            selected_file = file_dialog.selectedFiles()[0]
            original_image_data = read_image(selected_file)
            new_width, new_height = 200, 300
            resized_image_data = resize_image(original_image_data, new_width, new_height)
            pixmap = QPixmap()
            pixmap.loadFromData(resized_image_data)
            Rgtr.OQL_SignPhoto.setPixmap(pixmap)
            globals()["Blob_Image_Sign"] = resized_image_data

    @Exception_Handle
    def open_file_dialog_Emp_Nomi():
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        file_dialog = QFileDialog()
        file_dialog.setOptions(options)
        file_dialog.setNameFilter("Image Files (*.png *.jpg *.bmp)")
        file_dialog.setWindowTitle("Select an Image")
        file_dialog.setFileMode(QFileDialog.ExistingFile)
        if file_dialog.exec_() == QFileDialog.Accepted:
            selected_file = file_dialog.selectedFiles()[0]
            original_image_data = read_image(selected_file)
            new_width, new_height = 500, 300
            resized_image_data = resize_image(original_image_data, new_width, new_height)
            pixmap = QPixmap()
            pixmap.loadFromData(resized_image_data)
            Rgtr.OQL_NomPhoto.setPixmap(pixmap)
            globals()["Blob_Image_Nomi"] = resized_image_data

    @Exception_Handle
    def get_pixmap_from_blob(blob_data):
        # Convert the blob data to a QPixmap
        try:
            img = Image.open(BytesIO(blob_data))
            img = img.convert("RGB")  # Convert image mode if needed (PIL related)
            img.save("temp_image.jpg")  # Save the temporary image
            pixmap = QPixmap("temp_image.jpg")  # Load the temporary image as QPixmap
            return pixmap
        except:
            print("No BLOB")

    @Exception_Handle
    def Export_Emp_Data_XL():
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
        from openpyxl.utils.units import cm_to_EMU, pixels_to_EMU
        from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
        from openpyxl.styles import Alignment
        from openpyxl.styles import Font as xlfont
        c2e = cm_to_EMU
        p2e = pixels_to_EMU
        cellh = lambda x: c2e((x * 49.77) / 99)
        cellw = lambda x: c2e((x * (18.65 - 1.71)) / 10)
        db_data = DB_Fetch(
            "SELECT `emp_code`,`employee_name`,`gender`,`f_sp_name`,`date_of_birth`,`date_of_join`,`designation`,"
            "`phone_no`,`uan_no`,`esic_no`,`pan_no`,`aadhar_no`,`bank_account_no`,`bank_name`,"
            "`ifsc_code`,`address`, shift_1_salary,"
            " concat(shift_1_salary,',',shift_2_salary,',',shift_3_salary), "
            "`photo`,`signature`"
            "from register where active = 'Y' and emp_code like 'SIL%' ", False, "LOL")

        xl = openpyxl.load_workbook(fr'{ldir}\EXTERNAL\Register_Export.xlsx')
        xl.active = xl['Form A']
        xlc = xl.active
        rowc = 3
        colc = 2
        print(db_data[0])
        for step in db_data:
            colc = 2
            for i in step:
                xlc.cell(row=rowc, column=colc).value = i
                xlc.cell(row=rowc, column=colc).alignment = Alignment(horizontal='center', vertical='center')
                xlc.cell(row=rowc, column=colc).font = xlfont(name="Courier New", size=10)
                colc += 1
                if colc == 20:
                    break
            try:
                blob_data_photo = step[18]
                blob_data_sign = step[19]
                image_photo = Image.open(BytesIO(blob_data_photo))
                image_sign = Image.open(BytesIO(blob_data_sign))

                image_photo.save(fr'{ldir}\Temp\output_image_photo{rowc}.jpg')
                image_sign.save(fr'{ldir}\Temp\output_image_sign{rowc}.jpg')

                a = Image.open(fr'{ldir}\Temp\output_image_photo{rowc}.jpg')
                b = a.resize((80, 80))
                b.save(fr'{ldir}\Temp\output_image_photo{rowc}.jpg')
                img = openpyxl.drawing.image.Image(fr'{ldir}\Temp\output_image_photo{rowc}.jpg')
                size = XDRPositiveSize2D(p2e(img.width), p2e(img.height))
                column = colc - 1
                coloffset = cellw(0.3)
                row = rowc - 1
                rowoffset = cellh(0.6)
                marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
                img.anchor = OneCellAnchor(_from=marker, ext=size)
                xlc.add_image(img)
                # --
                a = Image.open(fr'{ldir}\Temp\output_image_sign{rowc}.jpg')
                b = a.resize((100, 40))
                b.save(fr'{ldir}\Temp\output_image_sign{rowc}.jpg')
                img = openpyxl.drawing.image.Image(fr'{ldir}\Temp\output_image_sign{rowc}.jpg')
                size = XDRPositiveSize2D(p2e(img.width), p2e(img.height))
                column = colc
                coloffset = cellw(0.6)
                row = rowc - 1
                rowoffset = cellh(2.0)
                marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
                img.anchor = OneCellAnchor(_from=marker, ext=size)
                xlc.add_image(img)
            except Exception as e:
                traceback.print_exc()
                pass
            rowc += 1
        xl.save(fr'{ldir}\Temp\Emp_Exp{Cur_Date_NF}.xlsx')
        subprocess.run(['start', 'excel', fr'{ldir}\Temp\Emp_Exp{Cur_Date_NF}.xlsx'], shell=True, check=True)
        UI_Confirmation(UI_Confirm_Win, "Excel Fetch is Completed Successfully")

    # _________ Functionality_List _________
    Customer_List = DB_Fetch("select emp_code,team,employee_name,blood_group,phone_no from register where "
                             "active = 'Y'  and  emp_code like 'SIL%' order by UID",
                             False, "LOL")

    Adjust_Table_Width(Rgtr.OQTB_EmployeeList, [10, 10, 33, 5, 15])
    Push_Table_Values(Rgtr.OQTB_EmployeeList, Customer_List, False)

    Rgtr.NonPFTOPF.setVisible(False)
    Rgtr.OQTB_EmployeeList.cellClicked.connect(Fetch_tbl)
    Rgtr.IQPB_Update.clicked.connect(lambda: Update_Emp_Details())
    Rgtr.GB_UpdateEmp.clicked.connect(lambda: Enable_Update())
    Rgtr.GB_Add_Emp.clicked.connect(lambda: Enable_Add())
    Rgtr.IQPB_Add.clicked.connect(lambda: Add_Emp_Details())
    Rgtr.OQTB_EmployeeList.setContextMenuPolicy(Qt.CustomContextMenu)
    Rgtr.OQTB_EmployeeList.customContextMenuRequested.connect(Right_Click_Menu)
    Rgtr.IQRB_PF.toggled.connect(lambda: Rgtr.IQLE_EmpCode.setText(Emp_code_Gen(True)))
    Rgtr.IQRB_NonPF.toggled.connect(lambda: Rgtr.IQLE_EmpCode.setText(Emp_code_Gen(False)))
    Rgtr.IQPB_PhotoBrowse.clicked.connect(lambda: open_file_dialog_Emp_Img())
    Rgtr.OQL_Sign.clicked.connect(lambda: open_file_dialog_Emp_Sign())
    Rgtr.IQPB_NomPhotoBrowse.clicked.connect(lambda: open_file_dialog_Emp_Nomi())
    Rgtr.Temp_View.stateChanged.connect(lambda: Temp_View_pwd())
    Rgtr.Temp_View.stateChanged.connect(lambda: NonPF_To_PF_Convert_Flow())
    Rgtr.NonPFTOPF.clicked.connect(lambda :NonPF_To_PF_Convert())

    Rgtr.IQLE_Search.textChanged.connect(lambda: Filter_Table_Data(Rgtr.OQTB_EmployeeList,
                                                                   Rgtr.IQLE_Search.text(), Customer_List, 2))

    Rgtr.Form_Export_XL.clicked.connect(lambda: Export_Emp_Data_XL())

# ___StandAlone Running___
if SAR == True:
    Register_FN(Rgtr)
    Rgtr.show()
    sys.exit(app.exec_())
